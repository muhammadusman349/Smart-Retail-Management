from io import BytesIO
from django.core.files.uploadedfile import InMemoryUploadedFile
from conf.celery import app
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Table,
    TableStyle,
    Spacer,
)


@app.task
def generate_payment_excel_file(payment_id):
    from .models import Payment

    try:
        payment = Payment.objects.get(id=payment_id)
    except Payment.DoesNotExist:
        print(f"Payment with ID {payment_id} does not exist.")
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Payment Receipt"

    # Define styles
    title_font = Font(size=16, bold=True)
    header_font = Font(size=14, bold=True)
    header_fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Set the headers with styles (start from row 3)
    payment_header = [
        "Payment ID",
        "Order ID",
        "Cutomer Name",
        "Payment Method",
        "Amount Paid",
        "Convenience Fee",
        "Total Amount",
        "Date"
    ]

    # Add title
    title_cell = sheet.cell(row=1, column=2)
    title_cell.value = "Payment Receipt Report"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(payment_header) + 1)

    # Create headers
    for col_num, header in enumerate(payment_header, start=2):
        cell = sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border

    # Add data
    row_index = 4
    data = [
        payment.id,
        payment.order.id,
        payment.order.user.name,
        payment.payment_method,
        payment.paid_amount,
        payment.convenience_fee,
        payment.total_amount,
        payment.date.strftime("%m/%d/%Y"),
    ]

    for col_num, value in enumerate(data, start=2):
        cell = sheet.cell(row=row_index, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="center")

    # Set column widths for better readability
    for col_num in range(2, len(payment_header) + 2):
        sheet.column_dimensions[chr(64 + col_num)].width = 20

    # Save the Excel file
    buff = BytesIO()
    workbook.save(buff)
    buff.seek(0)
    file = InMemoryUploadedFile(buff, "xlsx", f"Payment_Receipt{payment.id}.xlsx", None, buff.tell(), None)
    payment.excel_file.save(f"Payment_Receipt_{payment.id}.xlsx", file)
    payment.save()


@app.task
def generate_payment_pdf_file(payment_id):
    from .models import Payment

    try:
        payment = Payment.objects.get(id=payment_id)
    except Payment.DoesNotExist:
        print(f"Payment with ID {payment_id} does not exist.")
        return

    buff = BytesIO()
    doc = SimpleDocTemplate(buff, pagesize=landscape(letter))
    elements = []

    # Set up stylesheet and title
    stylesheet = getSampleStyleSheet()
    title_style = stylesheet["Title"]
    title_style.alignment = 1

    # Add heading
    heading = Paragraph("Payment Receipt Report", title_style)
    elements.append(heading)
    elements.append(Spacer(1, 12))

    # Create the payment data table
    payment_data = [
        ["Payment ID", "Order ID", "Cutomer Name", "Payment Method", "Amount Paid", "Convenience Fee", "Total Amount", "Date"],
        [
            payment.id,
            payment.order.id,
            payment.order.user.name,
            payment.payment_method,
            payment.paid_amount,
            payment.convenience_fee,
            payment.total_amount,
            payment.date.strftime("%m/%d/%Y")
        ]
    ]

    # Create the table with better styling
    t1 = Table(payment_data, colWidths=[100] * len(payment_data[0]))

    # Define table style
    table_style = TableStyle(
        [
            # Header row style
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkgreen),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),

            # Data row style
            ("ALIGN", (0, 1), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

            # Alternating row background
            ("BACKGROUND", (0, 2), (-1, 2), colors.lightgrey),

            # Grid lines
            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            # Padding and spacing
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]
    )
    t1.setStyle(table_style)
    elements.append(t1)
    elements.append(Spacer(1, 24))
    doc.build(elements)
    buff.seek(0)

    # Save the PDF file
    file = InMemoryUploadedFile(buff, "pdf", f"{payment.id}.pdf", None, buff.tell(), None)
    payment.pdf_file.save(f"Payment_Receipt_{payment.id}.pdf", file)
    payment.save()
